import openpyxl, json, sys
files = [
    "DTPS Leave Balance Sheet FY 26-27.xlsx",
    "MWU LEAVE BALANCE SHEET -FY 26-27.xlsx",
]
for f in files:
    print("="*80)
    print("FILE:", f)
    wb = openpyxl.load_workbook(f, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n-- Sheet: {sn}  (rows={ws.max_row}, cols={ws.max_column}) --")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=True):
            print(row)
